from utils import *
import pandas as pd
from openpyxl.chart import LineChart,Reference,Series
from openpyxl import load_workbook


if __name__ == "__main__":
    start_time = time.time()
    params = ['strategy','foresight','learning rate','hyper','cool']
    report_filename = 'report_ev.xlsx'
    executor_filename = 'executor_report.xls'
    steps_sheetname = 'steps'
    loss_sheetname = 'loss'
    rev_loss_sheetname = 'reverse_loss'
    data = {'executor': [*['' for k in params],*[i for i in list(np.loadtxt(executor_filename, 'int'))][:-1]]}
    loss, rev_loss = {}, {}
    
    with open("config.json", "r") as config_file: CONFIG = json.load(config_file)
    for i in CONFIG: 
        if i['target'] == 'prediction':
            data.update({i['prefix'] :   [*[i[k] for k in params],*list(np.loadtxt(i['report'], 'int', delimiter=',')[-len(data['executor'])+len(params):,0])]})
            loss.update({i['prefix'] : [*[i[k] for k in params],*list(np.loadtxt(i['report'], 'float', delimiter=',')[:,1])]})
            tmp = i['prefix']
        if i['target'] == 'reverse':
            rev_loss.update({i['prefix'] : [*[i[k] for k in params],*list(np.loadtxt(i['report'], 'float', delimiter=','))]})
            
    index = [*params, *[i for i in range(1,len(data['executor'])-len(params)+1)]]
    index1 = [*params, *[i for i in range(1,len(loss[tmp])-len(params)+1)]]
   
    with pd.ExcelWriter(report_filename) as writer:  
        df = pd.DataFrame(data,index=index)
        df.to_excel(writer,sheet_name = steps_sheetname)
        tf = pd.DataFrame(loss, index=index1)
        tf.to_excel(writer,sheet_name = loss_sheetname)
        rf = pd.DataFrame(rev_loss, index=index1)
        rf.to_excel(writer,sheet_name=rev_loss_sheetname)
        output('[Upd]'+report_filename)
        
    print('Steps to fall\n',df)
    print('Loss function for prediction\n',tf)
    print('Loss function for reverse\n',rf)
    #>>>charts   
    report = load_workbook(filename=report_filename)
    sheet = report.active
    sheet1 = report.get_sheet_by_name(loss_sheetname)
    chart = LineChart()
    k = 2
    for i in loss:
        values = Reference(sheet1, min_col = k, min_row = len(params) + 2, max_row = sheet.max_row)
        series = Series(values , title = i)
        chart.series.append(series)
        k += 1
    chart.title = " Absolute value of loss function (prediction)"
    chart.x_axis.title = " Iteration "
    chart.y_axis.title = " Loss function "
    sheet1.add_chart(chart, "K2") 
   
    sheet2 = report.get_sheet_by_name(rev_loss_sheetname)
    chart = LineChart()
    k = 2
    for i in rev_loss:
        values = Reference(sheet2, min_col = k, min_row = len(params) + 2, max_row = sheet.max_row)
        series = Series(values , title = i)
        chart.series.append(series)
        k += 1
    chart.title = " Absolute value of loss function (reverse)"
    chart.x_axis.title = " Iteration "
    chart.y_axis.title = " Loss function "
    sheet2.add_chart(chart, "H2") 
    report.save(report_filename) 
    output('[Upd]'+report_filename)
    output('Charts plotted.')

    output('Session of viewer.py ended in ','time',time.time()-start_time)  